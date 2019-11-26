##################   Merging files and Genbank file extracting

##when there is no libraries. Prefer to create "environments" to specific libraries (e.g. conda create environment)
##However, you can use as follows; installing libraries out of environments
import os,sys

try:
    import Bio
except ImportError:
    os.system('pip install biopython')
	
try:
    import Scipy
except ImportError:
    os.system('pip install scipy')
	
try:
    import numpy
except ImportError:
    os.system('pip install numpy')
	
try:
    import openpyxl
except ImportError:
    os.system('pip install openpyxl')
	
	
from Bio import SeqIO

import sys
from sys import argv

##################   concatenating gb files###

All_GBfiles = '{}{}_All_GBfiles.gb'.format(argv[2],argv[3]) #'F:/Python/room/All_GBfiles.gb' 


def concatFiles(): # concatenate txt files according to https://stackoverflow.com/questions/13613336/python-concatenate-text-files
    path = '{}'.format(argv[1])
    files = os.listdir(path)
    for idx, infile in enumerate(files):
        print ("File #" + str(idx) + "  " + infile)
    concat = ''.join([open(path + f).read() for f in files])
    with open(All_GBfiles, "w") as fo:
        fo.write(concat)

if __name__ == "__main__":
    concatFiles()

	
##################   extracting data from Genbank file###

from Bio import SeqIO
import os
input = All_GBfiles
          
output = '{}{}_Alldata.txt'.format(argv[2],argv[3]) #"F:/Python/room/Alldata.txt"

handle = open(input)

if not os.path.exists(output): #checks for a pre-existing file with the same name as the output
    for seq_record in SeqIO.parse(handle, "genbank"):
        for seq_feature in seq_record.features:
            if seq_feature.type=="source":
                try: #If you would like your script to run to completion even when that information is not present, you can wrap each access to seq_feature.qualifiers in a try-except block catching KeyError and IndexError
                    strain = seq_feature.qualifiers.get('strain')
                except (KeyError, IndexError):
                    strain = None
                try:
                    country = seq_feature.qualifiers.get('country')
                except (KeyError, IndexError):
                    country = None
                try:
                    isolation_source = seq_feature.qualifiers.get('isolation_source')
                except (KeyError, IndexError):
                    isolation_source = None
                try:
                    host = seq_feature.qualifiers.get('host')
                except (KeyError, IndexError):
                    host = None
                try:
                    phylum = seq_record.annotations["taxonomy"][3]
                except (KeyError, IndexError):
                    phylum = None
                try:
                    order = seq_record.annotations["taxonomy"][6]
                except (KeyError, IndexError):
                    order = None
                with open(output, "a") as ofile:
                     ofile.write(">{0}xyx{1}xyx{2}xyx{3}xyx{4}xyx{5}xyx{6}xyx{7}xyx{8}xyx{9}xyx{10}xyx{11}xyx{12}xyx{13}zyz\n".format(seq_record.annotations["organism"], 
                     phylum, order, seq_record.name, seq_record.description, strain, country, isolation_source, host, 
                     len(seq_record), seq_record.annotations['references'][0].authors, 
                     seq_record.annotations['references'][0].title, seq_record.annotations['references'][0].journal, 
                     seq_record.seq))
                    
else:
    print ("The output file already seem to exist in the current working directory {0}. Please change the name of the output file".format(os.getcwd())) #error msg

handle.close()

##################   Removing duplicate elements###

none_duplicate = '{}{}_no_duplicates.txt'.format(argv[2],argv[3])#"F:/Python/room/none_duplicate.txt"


import os,sys
 
def Remove(duplicate): # https://www.geeksforgeeks.org/python-remove-duplicates-list/
    final_list = [] 
    for data in duplicate: 
        if data not in final_list: 
            final_list.append(data) 
    return final_list

# Driver Code 
duplicate = open(output) #https://qiita.com/visualskyrim/items/1922429a07ca5f974467

sys.stdout=open(none_duplicate,"w") #https://stackoverflow.com/questions/7152762/how-to-redirect-print-output-to-a-file-using-python/38186276
print(Remove(duplicate))
#sys.stdout.close()
#duplicate.close()

##################   Adding first line for headernames###

f1 = open('{}{}_no_duplicates.txt'.format(argv[2],argv[3]), 'r')#open("F:/Python/room/none_duplicate.txt", 'r')
f2 = open('{}{}_no_duplicates_fw.txt'.format(argv[2],argv[3]), 'w')#open('F:/Python/room/none_duplicate_re.txt', 'w')#output final file


for j in f1:
   f2.write("\n" + j)


##################   Replace the target string###

f3 = open('{}{}_no_duplicates_fw.txt'.format(argv[2],argv[3]), 'r')#open('F:/Python/room/none_duplicate_re.txt', 'r')
f4 = open('{}{}_no_duplicates_fw2.txt'.format(argv[2],argv[3]), 'w')#open('F:/Python/room/none_duplicate_re2.txt', 'w')


for line in f3:
    f4.write(line.replace('xyx','\t').replace('zyz','\n').replace('n", ">','>').replace('[">','>')
	.replace('n"]','').replace("['",'').replace("']",'').replace("\>",'>').replace("n', ",'').replace('n", ','')
	.replace("\'>",'>').replace('\">','>').replace('\>','>').replace("[\'",'').replace('["','').replace("[\'",'')
	.replace("[","").replace("]","").replace("\'",''))


f3.close()
f4.close()


##################   header to xls###


import csv
import openpyxl
from openpyxl import Workbook

csv.field_size_limit()

csv.field_size_limit(100000000)

csv.field_size_limit()

##################   txt to xlsx###

input_file = '{}{}_no_duplicates_fw2.txt'.format(argv[2],argv[3])#'F:/Python/room/none_duplicate_re2.txt'
output_file = '{}{}_SpecimensList.xlsx'.format(argv[2],argv[3])#'F:/Python/room/none_duplicate_re2.xlsx'

headernames = ['Species', 'phylum', 'order', 'GBn', 'Description', 'Voucher', 'Country', 
'isolation_source', 'host', 'bp', 'authors', 'title', 'journal', 'sequence']

wb = openpyxl.Workbook()
ws = wb.worksheets[0]

with open(input_file, 'r') as data:
    reader = csv.reader(data, delimiter='\t')
    for row in reader:
        ws.append(row)
    for i, j in enumerate(headernames):
        ws.cell(row=1, column = i+1, value= j)

wb.save(output_file)

from openpyxl import load_workbook
workbook = load_workbook('{}{}_SpecimensList.xlsx'.format(argv[2],argv[3]))
sheet = workbook.active

sheet.delete_rows(idx=sheet.max_row)#https://medium.com/aubergine-solutions/working-with-excel-sheets-in-python-using-openpyxl-4f9fd32de87f

workbook.save('{}{}_SpecimensList.xlsx'.format(argv[2],argv[3]))

sys.stdout.close()
